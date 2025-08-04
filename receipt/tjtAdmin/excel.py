import openpyxl
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from django.conf import settings
import os
from pathlib import Path
# create function to create excel file

# get prexisting excel file
files = os.path.join(settings.MEDIA_ROOT, 'sheets')
file = [os.path.join(files, i) for i in os.listdir(files)][0]
file = os.path.abspath(file)

def xlsx(orderItems):
  wb = Workbook()
  ws = wb.active
  ws.title = 'TJT records'
  ws['A1']= 'DATE'
  ws['B1']= 'NAME'
  ws['C1']= 'CL'
  ws['D1']= 'DRINK'
  ws['E1']= 'QUANTITY'
  ws['F1']= 'PRICE'
  ws['G1']= 'TOTAL COST'

  # insert date in date column
  for orderitem in orderItems:
    x = orderitem
    cl, product =  stripProductName(x.product.name)
    new_row = (x.order.date.date(), x.order.client.name, cl , product, x.quantity, x.product.price, x.get_total)
    ws.append(new_row)

  # save file to in memory
  # stream
  return saveFile(wb)

def allTime(items):
  # get prexising excel file and load it
  wb = openpyxl.load_workbook(file)
  ws = wb['Orders 2025']

  # create and insert new rows which consist of all the order items
  for item in items:
    cl, product =  stripProductName(item.product.name)
    new_row = ('',item.order.date.date(), item.order.client.name,cl, product, item.quantity, item.product.price, item.get_total)
    ws.append(new_row)

  return saveFile(wb)

def preDB(orderItems):
  # get prexising excel file and load it
  wb = openpyxl.load_workbook(file)
  ws = wb['Orders 2025']

  # insert date in date column
  for orderitem in orderItems:
    x = orderitem
    cl, product =  stripProductName(x.product.name)
    new_row = (x.order.date.date(), x.order.client.name, cl , product, x.quantity, x.product.price, x.get_total)
    ws.append(new_row)
  
  return saveFile(wb)

def stripProductName(name):
  name_size = str(name).split(' ')
  size = name_size[-1]
  name_size.pop(-1)
  name_only = ""
  for item in name_size:
    name_only += f" {item}"
  return size, name_only.lstrip()

def saveFile(wb):
  # save to in-memory file
  stream = BytesIO()
  wb.save(stream)
  stream.seek(0)

  return stream
